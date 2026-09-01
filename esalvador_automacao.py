from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from datetime import datetime
import os
import re
import time
import shutil


# ============================================================
# CONFIGURAÇÕES
# ============================================================

URL_LOGIN = "https://esalvador.salvador.ba.gov.br"

URL_CONSULTA = "https://esalvador.salvador.ba.gov.br/consulta"

PASTA_PROGRAMA = os.path.dirname(os.path.abspath(__file__))

# ------------------------------------------------------------
# PLANILHA ORIGINAL
# ------------------------------------------------------------

ARQUIVO_ORIGINAL = os.path.join(
    PASTA_PROGRAMA,
    "GERAL.xlsx"
)

# ------------------------------------------------------------
# CÓPIA QUE SERÁ ATUALIZADA
# ------------------------------------------------------------

ARQUIVO_ATUALIZADO = os.path.join(
    PASTA_PROGRAMA,
    "GERAL_ATUALIZADO.xlsx"
)


# ============================================================
# COLUNAS DA PLANILHA
# ============================================================

COL_ORIGEM = 1
COL_PROCESSO = 2
COL_ANO = 3
COL_OBJETO = 4

# E = UNIDADE
COL_UNIDADE = 5

# F = DIAS NA UNIDADE
COL_DIAS = 6

# G = OBSERVAÇÃO
COL_OBSERVACAO = 7

LINHA_INICIAL = 2


# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================

def limpar_texto(texto):

    if texto is None:
        return ""

    texto = str(texto)

    texto = texto.replace("\xa0", " ")

    texto = re.sub(r"\s+", " ", texto)

    return texto.strip()


def normalizar_texto(texto):

    return limpar_texto(texto).lower()


def numero_limpo(valor):

    if valor is None:
        return ""

    texto = str(valor).strip()

    if texto.endswith(".0"):
        texto = texto[:-2]

    return texto


def extrair_numero(valor):

    if valor is None:
        return ""

    numeros = re.findall(
        r"\d+",
        str(valor)
    )

    if not numeros:
        return ""

    return numeros[-1]


# ============================================================
# VERIFICAR PLANILHA ORIGINAL
# ============================================================

def verificar_excel():

    print()
    print("=" * 70)
    print("VERIFICANDO PLANILHA")
    print("=" * 70)

    print()
    print("Planilha original:")

    print(
        ARQUIVO_ORIGINAL
    )

    if not os.path.exists(ARQUIVO_ORIGINAL):

        print()
        print("ERRO: GERAL.xlsx não foi encontrado.")

        print()
        print(
            "Coloque o arquivo GERAL.xlsx "
            "na mesma pasta do programa."
        )

        return False

    print()
    print("✓ GERAL.xlsx encontrado.")

    return True


# ============================================================
# CRIAR CÓPIA DA PLANILHA
# ============================================================

def criar_copia_planilha():

    try:

        if os.path.exists(
            ARQUIVO_ATUALIZADO
        ):

            try:

                os.remove(
                    ARQUIVO_ATUALIZADO
                )

            except PermissionError:

                print()
                print("=" * 70)
                print("ERRO: GERAL_ATUALIZADO.xlsx ESTÁ ABERTO")
                print("=" * 70)
                print()
                print(
                    "Feche o arquivo GERAL_ATUALIZADO.xlsx "
                    "e execute novamente."
                )

                return False

        shutil.copy2(
            ARQUIVO_ORIGINAL,
            ARQUIVO_ATUALIZADO
        )

        print()
        print(
            "✓ Cópia criada:"
        )

        print(
            ARQUIVO_ATUALIZADO
        )

        return True

    except Exception as erro:

        print()
        print(
            "ERRO AO CRIAR CÓPIA DA PLANILHA:"
        )

        print(erro)

        return False


# ============================================================
# ABRIR CÓPIA DA PLANILHA
# ============================================================

def abrir_planilha():

    try:

        print()
        print(
            "Abrindo GERAL_ATUALIZADO.xlsx..."
        )

        wb = load_workbook(
            ARQUIVO_ATUALIZADO
        )

        ws = wb.active

        print(
            f"✓ Aba utilizada: {ws.title}"
        )

        print(
            f"✓ Linhas encontradas: {ws.max_row}"
        )

        return wb, ws

    except PermissionError:

        print()
        print("=" * 70)
        print("ERRO: PLANILHA ABERTA")
        print("=" * 70)
        print()
        print(
            "Feche o GERAL_ATUALIZADO.xlsx "
            "antes de executar."
        )

        return None, None

    except Exception as erro:

        print()
        print(
            "Erro ao abrir planilha:"
        )

        print(erro)

        return None, None


# ============================================================
# SALVAR CÓPIA
# ============================================================

def salvar_planilha(wb):

    try:

        wb.save(
            ARQUIVO_ATUALIZADO
        )

        print()
        print(
            "✓ GERAL_ATUALIZADO.xlsx SALVO"
        )

        return True

    except PermissionError:

        print()
        print("=" * 70)
        print("ERRO AO SALVAR PLANILHA")
        print("=" * 70)
        print()
        print(
            "Feche o GERAL_ATUALIZADO.xlsx."
        )

        return False

    except Exception as erro:

        print()
        print(
            "Erro ao salvar:"
        )

        print(erro)

        return False


# ============================================================
# ABRIR PLANILHA AO FINAL
# ============================================================

def abrir_excel_no_final():

    try:

        print()
        print(
            "Abrindo GERAL_ATUALIZADO.xlsx..."
        )

        os.startfile(
            ARQUIVO_ATUALIZADO
        )

        print(
            "✓ Planilha atualizada aberta."
        )

    except Exception as erro:

        print()
        print(
            "Não foi possível abrir a planilha."
        )

        print(erro)


# ============================================================
# LOGIN
# ============================================================

def esperar_login(page):

    print()
    print("=" * 70)
    print("LOGIN DO eSALVADOR")
    print("=" * 70)

    print()
    print(
        "Faça o login manualmente no navegador."
    )

    print()
    print(
        "Depois de entrar no sistema,"
    )

    print(
        "volte para esta janela e pressione ENTER."
    )

    print()

    input(
        "Pressione ENTER para continuar..."
    )

    return True


# ============================================================
# ESPERAR PÁGINA ESTABILIZAR
# ============================================================

def esperar_pagina_estabilizar(page, timeout_extra=2500):

    try:

        page.wait_for_load_state(
            "networkidle",
            timeout=8000
        )

    except Exception:
        pass

    page.wait_for_timeout(
        timeout_extra
    )


# ============================================================
# ENTRAR NA CONSULTA
# ============================================================

def entrar_na_consulta(page):

    print()
    print("=" * 70)
    print("ABRINDO CONSULTA")
    print("=" * 70)

    try:

        page.goto(
            URL_CONSULTA,
            wait_until="domcontentloaded",
            timeout=60000
        )

        esperar_pagina_estabilizar(
            page,
            timeout_extra=2500
        )

    except Exception as erro:

        print()
        print(
            "Erro ao abrir consulta:"
        )

        print(erro)

    try:

        campo = page.get_by_text(
            "Nº INTERNO",
            exact=False
        ).first

        if campo.count() > 0:

            print(
                "✓ Tela CONSULTA aberta."
            )

            return True

    except Exception:
        pass

    try:

        processos = page.get_by_text(
            "PROCESSOS",
            exact=True
        ).first

        if processos.count() > 0:

            processos.click()

            page.wait_for_timeout(
                1000
            )

    except Exception:
        pass

    try:

        consulta = page.get_by_text(
            "CONSULTA",
            exact=True
        ).first

        if consulta.count() > 0:

            consulta.click()

            esperar_pagina_estabilizar(
                page,
                timeout_extra=2500
            )

    except Exception:
        pass

    try:

        campo = page.get_by_text(
            "Nº INTERNO",
            exact=False
        ).first

        if campo.count() > 0:

            print(
                "✓ Tela CONSULTA aberta."
            )

            return True

    except Exception:
        pass

    print()
    print(
        "ERRO: não foi possível abrir a CONSULTA."
    )

    return False


# ============================================================
# ENCONTRAR CAMPO DO PROCESSO
# ============================================================

def encontrar_input_processo(page):

    seletores = [

        "input[name*='interno' i]",

        "input[id*='interno' i]",

        "input[placeholder*='interno' i]",

        "input[name*='processo' i]",

        "input[id*='processo' i]",

        "input[placeholder*='processo' i]"
    ]

    for seletor in seletores:

        try:

            campos = page.locator(
                seletor
            )

            quantidade = campos.count()

            for i in range(
                quantidade
            ):

                campo = campos.nth(i)

                if campo.is_visible():

                    return campo

        except Exception:
            pass

    try:

        texto = page.get_by_text(
            "Nº INTERNO",
            exact=False
        ).first

        if texto.count() > 0:

            bloco = texto.locator(
                "xpath=.."
            )

            for _ in range(4):

                try:

                    campos = bloco.locator(
                        "input"
                    )

                    quantidade = campos.count()

                    for i in range(
                        quantidade
                    ):

                        campo = campos.nth(i)

                        if campo.is_visible():

                            return campo

                    bloco = bloco.locator(
                        "xpath=.."
                    )

                except Exception:
                    break

    except Exception:
        pass

    try:

        inputs = page.locator(
            "input"
        )

        quantidade = inputs.count()

        visiveis = []

        for i in range(
            quantidade
        ):

            campo = inputs.nth(i)

            try:

                if campo.is_visible():

                    visiveis.append(
                        campo
                    )

            except Exception:
                pass

        if len(visiveis) >= 1:

            return visiveis[0]

    except Exception:
        pass

    return None


# ============================================================
# ENCONTRAR CAMPO ANO
# ============================================================

def encontrar_input_ano(page):

    seletores = [

        "input[name*='ano' i]",

        "input[id*='ano' i]",

        "input[placeholder*='ano' i]"
    ]

    for seletor in seletores:

        try:

            campos = page.locator(
                seletor
            )

            quantidade = campos.count()

            for i in range(
                quantidade
            ):

                campo = campos.nth(i)

                if campo.is_visible():

                    return campo

        except Exception:
            pass

    try:

        inputs = page.locator(
            "input"
        )

        quantidade = inputs.count()

        visiveis = []

        for i in range(
            quantidade
        ):

            campo = inputs.nth(i)

            try:

                if campo.is_visible():

                    visiveis.append(
                        campo
                    )

            except Exception:
                pass

        if len(visiveis) >= 2:

            return visiveis[1]

    except Exception:
        pass

    return None


# ============================================================
# CLICAR BUSCAR
# ============================================================

def clicar_botao_buscar(page):

    try:

        botao = page.get_by_role(
            "button",
            name=re.compile(
                r"^\s*BUSCAR\s*$",
                re.IGNORECASE
            )
        ).first

        if botao.count() > 0:

            botao.click()

            print(
                "✓ BUSCAR clicado."
            )

            return True

    except Exception:
        pass

    try:

        botoes = page.locator(
            "button"
        )

        quantidade = botoes.count()

        for i in range(
            quantidade
        ):

            botao = botoes.nth(i)

            try:

                if not botao.is_visible():
                    continue

                texto = limpar_texto(
                    botao.inner_text()
                )

                if texto.upper() == "BUSCAR":

                    botao.click()

                    print(
                        "✓ BUSCAR clicado."
                    )

                    return True

            except Exception:
                pass

    except Exception:
        pass

    return False


# ============================================================
# ENCONTRAR A LINHA CORRETA NA TABELA DE RESULTADOS
# ============================================================

def linha_bate_com_processo(texto_linha, processo, ano):

    texto = normalizar_texto(
        texto_linha
    )

    padrao_processo = re.compile(
        r"(?<!\d)" + re.escape(str(processo)) + r"(?!\d)"
    )

    if not padrao_processo.search(texto):

        return False

    if ano:

        padrao_ano = re.compile(
            r"(?<!\d)" + re.escape(str(ano)) + r"(?!\d)"
        )

        if not padrao_ano.search(texto):

            return False

    return True


def encontrar_linha_da_tabela(page, processo, ano):

    try:

        linhas = page.locator(
            "table tr"
        )

        quantidade = linhas.count()

    except Exception:

        return None

    for i in range(
        quantidade
    ):

        linha = linhas.nth(i)

        try:

            if not linha.is_visible():
                continue

            texto_linha = linha.inner_text()

        except Exception:
            continue

        if linha_bate_com_processo(
            texto_linha,
            processo,
            ano
        ):

            return linha

    return None


# ============================================================
# CLICAR LUPA DE AÇÃO
# ============================================================

def clicar_lupa_da_linha(
    page,
    processo,
    ano
):

    print()
    print(
        f"Procurando processo {processo}/{ano}..."
    )

    try:

        page.wait_for_timeout(
            1500
        )

    except Exception:
        pass

    linha = encontrar_linha_da_tabela(
        page,
        processo,
        ano
    )

    if linha is None:

        print()
        print(
            "ERRO: processo não apareceu na tabela "
            "(ou não bateu processo+ano juntos)."
        )

        return False

    print(
        "✓ Linha do processo encontrada."
    )

    try:

        celulas = linha.locator(
            "td"
        )

        quantidade = celulas.count()

        print(
            f"Células encontradas: {quantidade}"
        )

        if quantidade == 0:

            return False

    except Exception:
        return False

    coluna_acoes = celulas.nth(
        quantidade - 1
    )

    print(
        "✓ Coluna AÇÕES localizada."
    )

    seletores = [

        "i.fa-search",

        ".fa-search",

        "i.glyphicon-search",

        ".glyphicon-search",

        "[title*='detalhar' i]",

        "[title*='visualizar' i]",

        "[aria-label*='detalhar' i]",

        "[aria-label*='visualizar' i]"
    ]

    for seletor in seletores:

        try:

            elementos = coluna_acoes.locator(
                seletor
            )

            quantidade_elementos = elementos.count()

            for i in range(
                quantidade_elementos
            ):

                elemento = elementos.nth(i)

                if not elemento.is_visible():
                    continue

                print(
                    "✓ Lupa de AÇÃO encontrada."
                )

                try:

                    pai = elemento.locator(
                        "xpath=ancestor::a[1]"
                    )

                    if pai.count() > 0:

                        pai.click()

                    else:

                        elemento.click()

                except Exception:

                    elemento.click()

                esperar_pagina_estabilizar(
                    page,
                    timeout_extra=2500
                )

                return True

        except Exception:
            pass

    try:

        elementos = coluna_acoes.locator(
            "a, button"
        )

        quantidade_elementos = elementos.count()

        for i in range(
            quantidade_elementos
        ):

            elemento = elementos.nth(i)

            try:

                if not elemento.is_visible():
                    continue

                print(
                    f"Clicando ação {i + 1}..."
                )

                elemento.click()

                esperar_pagina_estabilizar(
                    page,
                    timeout_extra=2500
                )

                return True

            except Exception:
                pass

    except Exception:
        pass

    print()
    print(
        "ERRO: não conseguiu clicar na lupa de AÇÃO."
    )

    return False


# ============================================================
# VALIDAR QUE A PÁGINA DE DETALHE É DO PROCESSO CERTO
# ============================================================

def validar_pagina_do_processo(page, processo, ano):

    try:

        texto = page.locator(
            "body"
        ).inner_text()

    except Exception:

        return False

    return linha_bate_com_processo(
        texto,
        processo,
        ano
    )


# ============================================================
# EXTRAIR SOMENTE UNIDADE E DIAS
# ------------------------------------------------------------
# CORREÇÃO PRINCIPAL (nova versão):
#
# O método antigo subia a partir do título "Localização Atual"
# até um ancestral comum e dependia de splitlines() para separar
# rótulo/valor. Isso falhava quando havia muitos <div> de
# wrapper entre o título e os campos, ou quando rótulo e valor
# ficavam na MESMA linha de texto (ex: "UnidadeATES").
#
# Agora a busca é ancorada em "Dias na Unidade", que é um texto
# único na página inteira. A partir dele, sobe até achar o
# elemento PAI que tem os campos irmãos (Órgão | Unidade |
# Dias na Unidade) e lê cada um separadamente, sem depender de
# quantos níveis de wrapper existem nem de quebras de linha.
#
# Mantém como fallback o método antigo (regex no texto inteiro
# da página) caso o método novo não encontre nada.
# ============================================================

def extrair_unidade_e_dias(page):

    print()
    print("=" * 70)
    print("LENDO LOCALIZAÇÃO ATUAL")
    print("=" * 70)

    esperar_pagina_estabilizar(
        page,
        timeout_extra=1500
    )

    unidade = ""
    dias = ""

    # --------------------------------------------------------
    # MÉTODO NOVO: ancora em "Dias na Unidade" e lê os campos
    # irmãos (Órgão / Unidade / Dias na Unidade) na mesma linha.
    # --------------------------------------------------------

    try:

        campo_dias = page.get_by_text(
            "Dias na Unidade",
            exact=True
        ).first

        if campo_dias.count() > 0:

            print(
                "✓ Texto 'Dias na Unidade' encontrado."
            )

            box_dias = campo_dias

            for nivel in range(1, 7):

                try:

                    box_dias = box_dias.locator(
                        "xpath=.."
                    )

                    filhos = box_dias.locator(
                        "xpath=./*"
                    )

                    qtd_filhos = filhos.count()

                except Exception:
                    continue

                if qtd_filhos < 2:
                    continue

                textos_filhos = []

                for i in range(qtd_filhos):

                    try:

                        texto_filho = limpar_texto(
                            filhos.nth(i).inner_text()
                        )

                    except Exception:

                        texto_filho = ""

                    textos_filhos.append(
                        texto_filho
                    )

                juntos = " | ".join(textos_filhos)

                print(
                    f"DEBUG nível {nivel}: {juntos}"
                )

                tem_unidade = re.search(
                    r"\bUnidade\b",
                    juntos,
                    re.IGNORECASE
                )

                tem_dias = re.search(
                    r"Dias\s+na\s+Unidade",
                    juntos,
                    re.IGNORECASE
                )

                if tem_unidade and tem_dias:

                    print(
                        f"✓ Bloco irmão encontrado (nível {nivel})."
                    )

                    for texto_filho in textos_filhos:

                        norm = normalizar_texto(
                            texto_filho
                        )

                        if norm.startswith("dias na unidade"):

                            valor = texto_filho[
                                len("Dias na Unidade"):
                            ].strip()

                            numero = extrair_numero(
                                valor
                            )

                            if numero:
                                dias = numero

                        elif (
                            norm.startswith("unidade")
                            and "dias" not in norm
                        ):

                            valor = texto_filho[
                                len("Unidade"):
                            ].strip()

                            if (
                                valor
                                and normalizar_texto(valor)
                                not in [
                                    "selecione",
                                    "selecione...",
                                    "selecione uma unidade",
                                    "-"
                                ]
                            ):

                                unidade = valor

                    if unidade or dias:
                        break

    except Exception as erro:

        print(
            "Erro no método novo (Dias na Unidade):"
        )

        print(erro)

    # --------------------------------------------------------
    # MÉTODO ANTIGO (fallback 1): subir a partir do título
    # "Localização Atual" e usar splitlines().
    # --------------------------------------------------------

    if not unidade or not dias:

        try:

            titulo = page.get_by_text(
                "Localização Atual",
                exact=True
            ).first

            if titulo.count() > 0:

                bloco = titulo

                for nivel in range(1, 8):

                    try:

                        bloco = bloco.locator(
                            "xpath=.."
                        )

                        texto_bloco = bloco.inner_text()

                        texto_bloco = limpar_texto(
                            texto_bloco
                        )

                        if (
                            re.search(
                                r"\bUnidade\b",
                                texto_bloco,
                                re.IGNORECASE
                            )
                            and
                            re.search(
                                r"Dias\s+na\s+Unidade",
                                texto_bloco,
                                re.IGNORECASE
                            )
                        ):

                            linhas = bloco.inner_text().splitlines()

                            linhas = [
                                limpar_texto(x)
                                for x in linhas
                                if limpar_texto(x)
                            ]

                            for i, linha in enumerate(linhas):

                                texto_normalizado = normalizar_texto(
                                    linha
                                )

                                if texto_normalizado == "unidade":

                                    if i + 1 < len(linhas):

                                        valor = limpar_texto(
                                            linhas[i + 1]
                                        )

                                        if (
                                            valor
                                            and
                                            normalizar_texto(valor)
                                            not in [
                                                "dias na unidade",
                                                "órgão",
                                                "unidade"
                                            ]
                                        ):

                                            if not unidade:
                                                unidade = valor

                                if (
                                    texto_normalizado
                                    == "dias na unidade"
                                ):

                                    if i + 1 < len(linhas):

                                        valor = limpar_texto(
                                            linhas[i + 1]
                                        )

                                        numero = extrair_numero(
                                            valor
                                        )

                                        if numero and not dias:

                                            dias = numero

                            if unidade or dias:

                                break

                    except Exception:
                        pass

        except Exception:
            pass

    # --------------------------------------------------------
    # MÉTODO ANTIGO (fallback 2): regex no texto inteiro do body.
    # --------------------------------------------------------

    if not unidade or not dias:

        try:

            texto = page.locator(
                "body"
            ).inner_text()

            texto = limpar_texto(
                texto
            )

            if not unidade:

                resultado = re.search(
                    r"Localização\s+Atual.*?"
                    r"\bUnidade\b\s*"
                    r"(?!Dias\s+na\s+Unidade)"
                    r"(.+?)"
                    r"Dias\s+na\s+Unidade",
                    texto,
                    re.IGNORECASE
                )

                if resultado:

                    valor = limpar_texto(
                        resultado.group(1)
                    )

                    valor = re.sub(
                        r"\bÓrgão\b.*",
                        "",
                        valor,
                        flags=re.IGNORECASE
                    )

                    unidade = limpar_texto(
                        valor
                    )

            if not dias:

                resultado = re.search(
                    r"Dias\s+na\s+Unidade\s*"
                    r"(\d+)",
                    texto,
                    re.IGNORECASE
                )

                if resultado:

                    dias = resultado.group(1)

        except Exception:
            pass

    unidade = limpar_texto(
        unidade
    )

    dias = extrair_numero(
        dias
    )

    if normalizar_texto(
        unidade
    ) in [
        "selecione",
        "selecione...",
        "selecione uma unidade",
        "-"
    ]:

        unidade = ""

    print()
    print("-" * 70)

    print(
        f"UNIDADE ENCONTRADA: {unidade}"
    )

    print(
        f"DIAS NA UNIDADE: {dias}"
    )

    print("-" * 70)

    if not unidade:

        print()
        print(
            "⚠ Não foi possível identificar a UNIDADE."
        )

    if not dias:

        print()
        print(
            "⚠ Não foi possível identificar os DIAS."
        )

    if not unidade and not dias:

        return None

    return {
        "unidade": unidade,
        "dias": dias,
        "data": datetime.now().strftime(
            "%d/%m/%Y %H:%M"
        )
    }


# ============================================================
# ATUALIZAR LINHA
# ============================================================

def atualizar_linha(
    ws,
    linha,
    resultado
):

    unidade_nova = limpar_texto(
        resultado.get(
            "unidade",
            ""
        )
    )

    dias_novo = extrair_numero(
        resultado.get(
            "dias",
            ""
        )
    )

    data = resultado.get(
        "data",
        ""
    )

    unidade_antiga = limpar_texto(
        ws.cell(
            linha,
            COL_UNIDADE
        ).value
    )

    dias_antigo = extrair_numero(
        ws.cell(
            linha,
            COL_DIAS
        ).value
    )

    print()
    print("=" * 70)
    print(
        f"ATUALIZANDO LINHA {linha}"
    )
    print("=" * 70)

    print()
    print(
        f"Unidade no Excel: {unidade_antiga}"
    )

    print(
        f"Unidade eSalvador: {unidade_nova}"
    )

    print()
    print(
        f"Dias no Excel: {dias_antigo}"
    )

    print(
        f"Dias eSalvador: {dias_novo}"
    )

    alteracoes = []

    if unidade_nova:

        if normalizar_texto(
            unidade_antiga
        ) != normalizar_texto(
            unidade_nova
        ):

            print()
            print(
                ">>> UNIDADE DIFERENTE <<<"
            )

            print(
                f"ANTIGA: {unidade_antiga}"
            )

            print(
                f"NOVA:   {unidade_nova}"
            )

            ws.cell(
                linha,
                COL_UNIDADE
            ).value = unidade_nova

            alteracoes.append(
                f"Unidade: "
                f"{unidade_antiga} -> "
                f"{unidade_nova}"
            )

        else:

            print()
            print(
                "✓ Unidade já está correta."
            )

    if dias_novo:

        if dias_novo != dias_antigo:

            print()
            print(
                ">>> DIAS DIFERENTES <<<"
            )

            print(
                f"ANTIGO: {dias_antigo}"
            )

            print(
                f"NOVO:   {dias_novo}"
            )

            ws.cell(
                linha,
                COL_DIAS
            ).value = int(
                dias_novo
            )

            alteracoes.append(
                f"Dias: "
                f"{dias_antigo} -> "
                f"{dias_novo}"
            )

        else:

            print()
            print(
                "✓ Dias já estão corretos."
            )

    if alteracoes:

        observacao_antiga = limpar_texto(
            ws.cell(
                linha,
                COL_OBSERVACAO
            ).value
        )

        registro = (
            f"[{data}] "
            + " | ".join(
                alteracoes
            )
        )

        if observacao_antiga:

            ws.cell(
                linha,
                COL_OBSERVACAO
            ).value = (
                observacao_antiga
                + " | "
                + registro
            )

        else:

            ws.cell(
                linha,
                COL_OBSERVACAO
            ).value = registro

        print()
        print(
            "✓ LINHA ATUALIZADA."
        )

        return True

    print()
    print(
        "✓ Nenhuma alteração necessária."
    )

    return False


# ============================================================
# VOLTAR PARA CONSULTA
# ============================================================

def voltar_para_consulta(page):

    try:

        page.goto(
            URL_CONSULTA,
            wait_until="domcontentloaded",
            timeout=60000
        )

        esperar_pagina_estabilizar(
            page,
            timeout_extra=1500
        )

        return True

    except Exception as erro:

        print()
        print(
            "Erro ao voltar para consulta:"
        )

        print(erro)

        return False


# ============================================================
# PROCESSAR
# ============================================================

def processar():

    if not verificar_excel():

        return

    if not criar_copia_planilha():

        return

    wb, ws = abrir_planilha()

    if wb is None:

        return

    with sync_playwright() as p:

        print()
        print("=" * 70)
        print("ABRINDO eSALVADOR")
        print("=" * 70)

        navegador = p.chromium.launch(
            headless=False
        )

        page = navegador.new_page(
            viewport={
                "width": 1536,
                "height": 864
            }
        )

        try:

            page.goto(
                URL_LOGIN,
                wait_until="domcontentloaded",
                timeout=60000
            )

        except Exception as erro:

            print()
            print(
                "ERRO AO ABRIR eSALVADOR:"
            )

            print(erro)

            navegador.close()

            wb.close()

            return

        esperar_login(
            page
        )

        if not entrar_na_consulta(
            page
        ):

            navegador.close()

            wb.close()

            return

        total = 0
        atualizados = 0
        corretos = 0
        erros = 0

        ultima_linha = ws.max_row

        for linha in range(
            LINHA_INICIAL,
            ultima_linha + 1
        ):

            processo = numero_limpo(
                ws.cell(
                    linha,
                    COL_PROCESSO
                ).value
            )

            ano = numero_limpo(
                ws.cell(
                    linha,
                    COL_ANO
                ).value
            )

            if not processo:

                continue

            total += 1

            print()
            print()
            print("#" * 70)

            print(
                f"PROCESSO {total}"
            )

            print(
                f"Linha Excel: {linha}"
            )

            print(
                f"Processo: {processo}/{ano}"
            )

            print("#" * 70)

            if not entrar_na_consulta(
                page
            ):

                erros += 1

                continue

            try:

                limpar = page.get_by_role(
                    "button",
                    name=re.compile(
                        r"^\s*LIMPAR\s*$",
                        re.IGNORECASE
                    )
                ).first

                if limpar.count() > 0:

                    if limpar.is_visible():

                        limpar.click()

                        page.wait_for_timeout(
                            500
                        )

            except Exception:
                pass

            campo_processo = encontrar_input_processo(
                page
            )

            if campo_processo is None:

                print()
                print(
                    "ERRO: CAMPO PROCESSO NÃO ENCONTRADO."
                )

                erros += 1

                continue

            print(
                "✓ Campo Nº INTERNO encontrado."
            )

            campo_processo.fill(
                processo
            )

            campo_ano = encontrar_input_ano(
                page
            )

            if campo_ano is not None:

                campo_ano.fill(
                    ano
                )

                print(
                    f"✓ Ano preenchido: {ano}"
                )

            else:

                print(
                    "⚠ Campo ANO não encontrado."
                )

            if not clicar_botao_buscar(
                page
            ):

                print(
                    "ERRO: botão BUSCAR não encontrado."
                )

                erros += 1

                continue

            esperar_pagina_estabilizar(
                page,
                timeout_extra=2000
            )

            if not clicar_lupa_da_linha(
                page,
                processo,
                ano
            ):

                erros += 1

                continue

            if not validar_pagina_do_processo(
                page,
                processo,
                ano
            ):

                print()
                print(
                    "ERRO: a página de detalhe aberta não "
                    "corresponde ao processo/ano esperado — "
                    "pulando esta linha para não gravar dado "
                    "errado."
                )

                erros += 1

                voltar_para_consulta(
                    page
                )

                continue

            resultado = extrair_unidade_e_dias(
                page
            )

            if resultado is None:

                print()
                print(
                    "ERRO: não foi possível extrair "
                    "UNIDADE/DIAS."
                )

                erros += 1

                voltar_para_consulta(
                    page
                )

                continue

            mudou = atualizar_linha(
                ws,
                linha,
                resultado
            )

            if mudou:

                atualizados += 1

            else:

                corretos += 1

            if not salvar_planilha(
                wb
            ):

                print()
                print(
                    "ERRO AO SALVAR."
                )

                print(
                    "Processamento interrompido."
                )

                erros += 1

                break

            voltar_para_consulta(
                page
            )

            time.sleep(
                1
            )

        print()
        print()
        print("=" * 70)
        print("PROCESSAMENTO CONCLUÍDO")
        print("=" * 70)

        print()
        print(
            f"Processos analisados: {total}"
        )

        print(
            f"Processos atualizados: {atualizados}"
        )

        print(
            f"Processos já corretos: {corretos}"
        )

        print(
            f"Erros: {erros}"
        )

        print()

        salvar_planilha(
            wb
        )

        navegador.close()

    try:

        wb.close()

    except Exception:
        pass

    abrir_excel_no_final()


# ============================================================
# INICIAR
# ============================================================

if __name__ == "__main__":

    print()
    print("=" * 70)
    print("ATUALIZADOR DE TRAMITAÇÃO - eSALVADOR")
    print("=" * 70)

    print()
    print(
        "Arquivo original:"
    )

    print(
        ARQUIVO_ORIGINAL
    )

    print()
    print(
        "Arquivo atualizado:"
    )

    print(
        ARQUIVO_ATUALIZADO
    )

    print()

    try:

        processar()

    except Exception as erro:

        print()
        print("=" * 70)
        print("ERRO INESPERADO")
        print("=" * 70)
        print()
        print(erro)

        import traceback

        traceback.print_exc()

    print()
    print("=" * 70)
    print("PROGRAMA FINALIZADO")
    print("=" * 70)

    input(
        "Pressione ENTER para fechar..."
    )