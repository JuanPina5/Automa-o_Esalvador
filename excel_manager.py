import os
import shutil
from openpyxl import load_workbook
import unicodedata

def normalize_header(header):
    if not header:
        return ""
    text = str(header).strip().lower()
    text = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('utf-8')
    return text.replace(" ", "")

class ExcelManager:
    def __init__(self, original_path, updated_path, logger):
        self.original_path = original_path
        self.updated_path = updated_path
        self.logger = logger
        self.wb = None
        self.ws = None
        self.col_map = {}
        self.header_row = 1

    def setup(self):
        self.logger.section("VERIFICANDO PLANILHA")
        if not os.path.exists(self.original_path):
            self.logger.error(f"Arquivo não encontrado: {self.original_path}")
            return False

        try:
            shutil.copy2(self.original_path, self.updated_path)
            self.logger.success(f"Cópia criada: {self.updated_path}")
            self.wb = load_workbook(self.updated_path)
            self.ws = self.wb.active
            self.logger.success(f"Planilha aberta. Aba utilizada: {self.ws.title}")
            return self._map_columns()
        except PermissionError:
            self.logger.error(f"O arquivo {self.updated_path} está aberto. Feche-o e tente novamente.")
            return False
        except Exception as e:
            self.logger.error(f"Erro ao preparar o Excel: {e}")
            return False

    def _map_columns(self):
        # Nomes normalizados que vamos procurar
        required_headers = {
            "processo": ["processo", "nprocesso", "numprocesso", "numeroprocesso", "noprocesso", "nodoprocesso"],
            "ano": ["ano"],
            "localizacaoatual": ["localizacaoatual", "unidade", "localizacao"],
            "tramitacaosetor(dias)": ["tramitacaosetor(dias)", "diasnaunidade", "tramitacaosetor"]
        }
        
        header_row = None
        for row in range(1, 11):
            found_cols = 0
            temp_map = {}
            for col in range(1, self.ws.max_column + 1):
                val = self.ws.cell(row=row, column=col).value
                norm_val = normalize_header(val)
                
                # Verifica a qual chave esse cabeçalho pertence
                for key, variations in required_headers.items():
                    if norm_val in variations:
                        temp_map[key] = col
                        found_cols += 1
                        break
                        
            if found_cols >= 2:
                header_row = row
                self.col_map = temp_map
                break

        if not header_row:
            self.logger.error("Cabeçalhos obrigatórios não encontrados.")
            return False

        missing = [req for req in required_headers.keys() if req not in self.col_map]
        if missing:
            self.logger.error(f"Colunas faltando (ou com nomes diferentes): {missing}")
            return False

        self.header_row = header_row
        self.logger.success("Colunas mapeadas dinamicamente.")
        return True

    def get_processes(self):
        processes = []
        for row in range(self.header_row + 1, self.ws.max_row + 1):
            processo = self.ws.cell(row=row, column=self.col_map["processo"]).value
            ano = self.ws.cell(row=row, column=self.col_map["ano"]).value
            
            # Limpa '.0' de números caso venham como float
            processo_str = str(processo).strip() if processo else ""
            if processo_str.endswith(".0"): processo_str = processo_str[:-2]
            
            ano_str = str(ano).strip() if ano else ""
            if ano_str.endswith(".0"): ano_str = ano_str[:-2]
            
            if processo_str:
                processes.append({
                    "row": row,
                    "processo": processo_str,
                    "ano": ano_str
                })
        return processes

    def update_process(self, row, unidade, dias):
        col_unidade = self.col_map["localizacaoatual"]
        col_dias = self.col_map["tramitacaosetor(dias)"]

        val_unidade_antiga = self.ws.cell(row=row, column=col_unidade).value
        val_dias_antigo = self.ws.cell(row=row, column=col_dias).value

        mudou = False
        self.logger.info("Excel:")
        
        if unidade is not None:
            unidade_str = str(unidade).strip()
            antiga_str = str(val_unidade_antiga).strip() if val_unidade_antiga else ""
            if unidade_str.lower() != antiga_str.lower():
                self.ws.cell(row=row, column=col_unidade).value = unidade_str
                self.logger.info(f"Localização Atual: {antiga_str} → {unidade_str}")
                mudou = True
            else:
                self.logger.info(f"Localização Atual: Mantida ({antiga_str})")

        if dias is not None:
            try:
                dias_int = int(dias)
                antiga_dias = int(val_dias_antigo) if val_dias_antigo is not None else None
                if dias_int != antiga_dias:
                    self.ws.cell(row=row, column=col_dias).value = dias_int
                    self.logger.info(f"Tramitação Setor: {val_dias_antigo} → {dias_int}")
                    mudou = True
                else:
                    self.logger.info(f"Tramitação Setor: Mantido ({antiga_dias})")
            except ValueError:
                pass
        
        if mudou:
            self.logger.success("Planilha atualizada")
        return mudou

    def save(self):
        try:
            self.wb.save(self.updated_path)
            self.logger.success("Salvamento concluído")
            return True
        except Exception as e:
            self.logger.error(f"Erro ao salvar planilha: {e}")
            return False
